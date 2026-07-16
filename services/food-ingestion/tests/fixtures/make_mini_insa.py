"""Generate the minimal synthetic PortFIR workbook used by the test suite.

Mirrors the real BDCA v7.1-2026 structure (3 sheets in order, banner row,
header row copied verbatim from :mod:`food_ingestion.portfir.nutrient_map`,
53-column layout) closely enough to pass ``parser.py``'s structural
validation, but with a handful of fabricated food rows engineered to
exercise every scenario the test suite checks. Contains no real INSA data
-- safe to commit.

Run directly to (re)write ``tests/fixtures/mini_insa.xlsx``::

    python -m tests.fixtures.make_mini_insa
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from food_ingestion.portfir import nutrient_map as nm

FIXTURE_PATH = Path(__file__).with_name("mini_insa.xlsx")

# Each row: (cod, name, level1, level2, level3, {canonical_code: value, ...})
# The dict only needs to name the nutrients that are MEASURED for that row;
# everything else is left blank (MISSING), matching real-world sparsity.
Row = tuple[object, str, str, str, str | None, dict[str, float]]

ROWS: list[Row] = [
    # 1: baseline "raw" preparation-state, full-ish macro coverage.
    (
        "10001",
        "Feijão cru",
        "Produtos hortícolas e derivados",
        "Leguminosas frescas",
        "Feijão verde",
        {
            "ENERC": 31, "ENERC_KJ": 130, "FAT": 0.2, "CHOAVL": 4.5, "SUGAR": 2.0,
            "FIBTG": 3.2, "PROCNT": 2.0, "WATER": 90.0, "ASH": 0.6, "NA": 4, "K": 220,
        },
    ),
    # 2: "cooked" preparation-state, blank Level-3, blank micronutrients,
    # a genuine measured 0 (FAT and ALC).
    (
        "10002",
        "Cenoura cozida",
        "Produtos hortícolas e derivados",
        "Raízes e tubérculos",
        None,
        {
            "ENERC": 35, "ENERC_KJ": 146, "FAT": 0.0, "CHOAVL": 6.0, "SUGAR": 4.0,
            "FIBTG": 2.8, "PROCNT": 0.8, "WATER": 90.0, "ALC": 0.0,
        },
    ),
    # 3: alcoholic beverage -> per-100 ml basis; energy consistent.
    (
        "20001",
        "Vinho tinto",
        "Bebidas alcoólicas",
        "Vinhos",
        "Vinho tinto de mesa",
        {
            "ENERC": 70, "ENERC_KJ": 293, "CHOAVL": 0.3, "SUGAR": 0.3, "ALC": 10.0,
            "WATER": 88.0, "ASH": 0.3, "K": 100,
        },
    ),
    # 4: numeric (int) Cod, mirrors the one real int-Cod row (Cod 40200010).
    (
        30001,
        "Maçã crua",
        "Frutos e produtos derivados de frutos",
        "Frutos frescos",
        "Pomóideas",
        {
            "ENERC": 52, "ENERC_KJ": 218, "FAT": 0.2, "CHOAVL": 11.4, "SUGAR": 10.0,
            "FIBTG": 2.4, "PROCNT": 0.3, "WATER": 85.0,
        },
    ),
    # 5: energy-consistency WARNING (kJ far off kcal*4.184), published anyway.
    (
        "40001",
        "Bolacha maria",
        "Cereais e produtos à base de cereais",
        "Produtos de pastelaria",
        "Bolachas",
        {
            "ENERC": 400, "ENERC_KJ": 1000, "FAT": 12.0, "CHOAVL": 70.0, "SUGAR": 20.0,
            "FIBTG": 2.0, "PROCNT": 6.0, "WATER": 3.0,
        },
    ),
    # 6: blank food name -> missing_food_name ERROR -> quarantined.
    (
        "50001",
        "",
        "Temperos, molhos e condimentos",
        "Molhos",
        None,
        {"ENERC": 100, "CHOAVL": 20.0, "NA": 500},
    ),
    # 7: duplicate Cod (matches row 1) -> duplicate_cod ERROR -> both
    # quarantined. Also a trailing-period Level-1 variant of row 1's label
    # -> generic level1_label_variant WARNING (on every row sharing the
    # normalised group, including the still-published row 2).
    (
        "10001",
        "Feijão cru (duplicado)",
        "Produtos hortícolas e derivados.",
        "Leguminosas frescas",
        "Feijão verde",
        {"ENERC": 31, "CHOAVL": 4.5, "PROCNT": 2.0},
    ),
    # 8: negative value -> ERROR -> quarantined.
    (
        "60001",
        "Produto com valor negativo",
        "Temperos, molhos e condimentos",
        "Molhos",
        None,
        {"ENERC": 50, "CHOAVL": 5.0, "NA": -5},
    ),
    # 9: Level-3 near-duplicate of row 4's "Pomóideas" (case + whitespace)
    # -> level3_label_variant WARNING, still published.
    (
        "70001",
        "Pêra crua",
        "Frutos e produtos derivados de frutos",
        "Frutos frescos",
        "pomóideas ",
        {
            "ENERC": 48, "ENERC_KJ": 201, "FAT": 0.1, "CHOAVL": 10.5, "SUGAR": 9.0,
            "FIBTG": 3.0, "PROCNT": 0.3, "WATER": 84.0,
        },
    ),
    # 10: macro-sum WARNING (105 g < sum <= 120 g) -- normal analytical
    # variance for cured/aged products, still published (sum = 108 g).
    (
        "80001",
        "Queijo curado teste",
        "Leite e produtos lácteos",
        "Queijos curados",
        "Queijo curado",
        {
            "ENERC": 350, "ENERC_KJ": 1464, "FAT": 30.0, "CHOAVL": 2.0, "PROCNT": 25.0,
            "WATER": 48.0, "ASH": 3.0,
        },
    ),
    # 11: macro-sum ERROR (sum > 120 g, no single macro > 100 g on its own)
    # -> quarantined.
    (
        "90001",
        "Erro grosseiro de soma",
        "Pratos compostos",
        "Teste",
        None,
        {"FAT": 60.0, "CHOAVL": 50.0, "PROCNT": 20.0, "WATER": 10.0},
    ),
    # 12: single-macro ERROR (one nutrient alone > 100 g/100g, physically
    # impossible) -> quarantined. Sum stays <= 105 g so this isolates the
    # single-macro rule from the macro-sum rule.
    (
        "90002",
        "Erro de macro único",
        "Água e bebidas à base de água",
        "Teste",
        None,
        {"WATER": 101.0},
    ),
]


def _headers() -> list[str]:
    headers = [nm.NON_NUTRIENT_HEADERS[i] for i in range(1, 6)]
    headers.extend(mapping.source_label for mapping in nm.NUTRIENT_MAP.values())
    return headers


def build_workbook(rows: list[Row] | None = None) -> Workbook:
    """Build the synthetic workbook. Defaults to the full `ROWS` scenario
    set; callers (mainly tests) may pass a smaller/custom row list to build
    a workbook covering just one scenario in isolation.
    """
    rows = ROWS if rows is None else rows
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = nm.DATA_SHEET_NAME

    data_sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    data_sheet.cell(1, 3, "Grupo e subgrupos (synthetic fixture banner)")
    data_sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=nm.EXPECTED_COLUMN_COUNT)
    data_sheet.cell(1, 6, "Valores por 100 g / 100 ml (synthetic fixture banner)")

    for column_index, header in enumerate(_headers(), start=1):
        data_sheet.cell(nm.HEADER_ROW_INDEX, column_index, header)

    canonical_to_index: dict[str, int] = {
        mapping.canonical_code: nm._column_letter_to_index(mapping.source_column)
        for mapping in nm.NUTRIENT_MAP.values()
    }

    row_index = nm.FIRST_DATA_ROW_INDEX
    for cod, name, level1, level2, level3, values in rows:
        data_sheet.cell(row_index, 1, cod)
        data_sheet.cell(row_index, 2, name)
        data_sheet.cell(row_index, 3, level1)
        data_sheet.cell(row_index, 4, level2)
        if level3 is not None:
            data_sheet.cell(row_index, 5, level3)
        for canonical_code, value in values.items():
            data_sheet.cell(row_index, canonical_to_index[canonical_code], value)
        row_index += 1

    _write_legend_sheet(workbook)
    _write_info_sheet(workbook)
    return workbook


def _write_legend_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(nm.LEGEND_SHEET_NAME)
    sheet.append(["Componentes\n[Unidades]", "EUROFIRCT1", "INFDSTAG2"])
    for mapping in nm.NUTRIENT_MAP.values():
        sheet.append([mapping.source_label, mapping.eurofir_code, mapping.infoods_code])
    for declared in nm.DECLARED_BUT_ABSENT:
        sheet.append([declared.legend_label, declared.eurofir_code, declared.infoods_code])


def _write_info_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(nm.INFO_SHEET_NAME)
    sheet.append([nm.ATTRIBUTION + " (synthetic fixture -- not real INSA data)"])


def main() -> None:
    workbook = build_workbook()
    FIXTURE_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(FIXTURE_PATH)
    print(f"Wrote {FIXTURE_PATH}")


if __name__ == "__main__":
    main()
