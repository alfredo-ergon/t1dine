"""Structural parsing of the PortFIR (INSA BDCA v7.1-2026) workbook into
immutable staging records.

This module never writes to the source workbook -- it is opened with
``read_only=True`` and only ever read from. Structural drift from the
documented layout (wrong sheets, wrong headers, too few columns) raises
:class:`WorkbookStructureError`: fail closed rather than guess at a
different shape. Row-level data-quality problems (negative values,
implausible macros, duplicate codes, ...) are *not* raised here -- they are
classified by :mod:`food_ingestion.portfir.quality` so that one bad row
never blocks the other 1375.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from . import nutrient_map as nm


class WorkbookStructureError(Exception):
    """Raised when the workbook does not match the documented BDCA v7.1-2026
    structure this adapter was built against.
    """


@dataclass(frozen=True)
class RawNutrientValue:
    source_column: str
    source_label: str
    original_value: float | int | None
    original_unit: str
    value_state: str  # "MEASURED" | "MISSING"


@dataclass(frozen=True)
class PortfirSourceRecord:
    source_org: str
    source_dataset: str
    source_version: str
    source_record_id: str
    original_food_name: str
    foodex2_level1: str
    foodex2_level2: str
    foodex2_level3: str | None
    basis: str  # "g" | "ml"
    raw_nutrients: list[RawNutrientValue]
    import_timestamp: str
    importer_version: str
    source_file_sha256: str
    mapping_version: str
    licence_status: str
    attribution: str
    transformation_notes: list[str] = field(default_factory=list)


def sha256_file(path: str | Path) -> str:
    """SHA-256 hex digest of a file's bytes, read in fixed-size chunks so
    large workbooks do not need to be loaded into memory at once.
    """
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def coerce_cod(value: Any) -> str:
    """Coerce a ``Cod`` cell value to its canonical string form.

    Handles the one documented case in the real BDCA v7.1-2026 export where a
    numeric-looking food code is stored as an Excel number rather than text
    (openpyxl then returns a Python ``int`` instead of ``str``).
    """
    if isinstance(value, bool):
        raise ValueError(f"Unexpected boolean Cod value: {value!r}")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"Unexpected non-integer numeric Cod value: {value!r}")
        return str(int(value))
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    raise ValueError(f"Unexpected Cod value type: {type(value)!r} ({value!r})")


def _validate_structure(workbook: Any) -> None:
    actual_sheets = tuple(workbook.sheetnames)
    if actual_sheets != nm.EXPECTED_SHEET_NAMES:
        raise WorkbookStructureError(
            "Unexpected sheet names/order. Expected "
            f"{nm.EXPECTED_SHEET_NAMES!r}, got {actual_sheets!r}."
        )

    worksheet = workbook[nm.DATA_SHEET_NAME]
    if worksheet.max_column is None or worksheet.max_column < nm.EXPECTED_COLUMN_COUNT:
        raise WorkbookStructureError(
            f"Data sheet has {worksheet.max_column} columns; expected at "
            f"least {nm.EXPECTED_COLUMN_COUNT} (A..BA)."
        )
    if worksheet.max_row is None or worksheet.max_row < nm.FIRST_DATA_ROW_INDEX:
        raise WorkbookStructureError(
            f"Data sheet has {worksheet.max_row} rows; expected at least "
            f"{nm.FIRST_DATA_ROW_INDEX} (header + >=1 data row)."
        )

    for column_index in range(1, nm.EXPECTED_COLUMN_COUNT + 1):
        actual = nm.normalize_header(worksheet.cell(nm.HEADER_ROW_INDEX, column_index).value)
        expected = nm.normalize_header(nm.expected_header_text(column_index))
        if actual != expected:
            raise WorkbookStructureError(
                f"Header drift at column {column_index}: expected {expected!r}, "
                f"got {actual!r}."
            )


def parse_records(path: str | Path, now: str | None = None) -> list[PortfirSourceRecord]:
    """Parse every food row of the PortFIR BDCA data sheet into an immutable
    staging record.

    Raw values are preserved verbatim: a blank nutrient cell becomes
    ``value_state="MISSING"`` with ``original_value=None`` (never inferred as
    zero); a genuine numeric ``0`` becomes ``value_state="MEASURED"``.

    ``now`` becomes ``import_timestamp`` on every record; when omitted it
    defaults to the fixed constant :data:`nutrient_map.DEFAULT_NOW` so
    determinism tests never depend on wall-clock time. ``import_timestamp``
    must stay out of any determinism/equality comparison of the *canonical*
    records -- see :mod:`food_ingestion.portfir.canonical`.
    """
    path = Path(path)
    file_sha256 = sha256_file(path)
    effective_now = now if now is not None else nm.DEFAULT_NOW

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        _validate_structure(workbook)
        worksheet = workbook[nm.DATA_SHEET_NAME]
        records: list[PortfirSourceRecord] = []
        # openpyxl's read-only mode returns a lightweight `EmptyCell`
        # placeholder (no `.column_letter`/`.row`/`.coordinate`) for cells
        # entirely absent from a row's underlying XML, alongside normal
        # `ReadOnlyCell` objects that merely hold a blank/empty value -- both
        # occur in the real workbook. Column letters are therefore taken
        # positionally from `nm.NUTRIENT_COLUMNS` and the row number from an
        # explicit counter, never from cell attributes.
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=nm.FIRST_DATA_ROW_INDEX, max_col=nm.EXPECTED_COLUMN_COUNT),
            start=nm.FIRST_DATA_ROW_INDEX,
        ):
            cod_cell, name_cell, level1_cell, level2_cell, level3_cell, *nutrient_cells = row
            if cod_cell.value is None and name_cell.value is None:
                # Trailing blank row padding at the end of the used range --
                # not a data anomaly worth quarantining.
                continue

            notes: list[str] = []
            raw_cod = cod_cell.value
            cod = coerce_cod(raw_cod)
            if isinstance(raw_cod, (int, float)) and not isinstance(raw_cod, bool):
                notes.append(
                    f"Cod coerced from numeric cell ({raw_cod!r}) to string ({cod!r})."
                )

            level1_raw = level1_cell.value
            level1_for_basis = (level1_raw or "").strip()
            basis = "ml" if level1_for_basis == nm.ALCOHOLIC_BEVERAGES_LEVEL1 else "g"

            raw_nutrients: list[RawNutrientValue] = [
                _parse_nutrient_value(column_letter, cell.value, row_index)
                for column_letter, cell in zip(nm.NUTRIENT_COLUMNS, nutrient_cells)
            ]

            records.append(
                PortfirSourceRecord(
                    source_org=nm.SOURCE_ORG,
                    source_dataset=nm.SOURCE_DATASET,
                    source_version=nm.SOURCE_VERSION,
                    source_record_id=cod,
                    original_food_name=name_cell.value if name_cell.value is not None else "",
                    foodex2_level1=level1_raw if level1_raw is not None else "",
                    foodex2_level2=(
                        level2_cell.value if level2_cell.value is not None else ""
                    ),
                    foodex2_level3=level3_cell.value,
                    basis=basis,
                    raw_nutrients=raw_nutrients,
                    import_timestamp=effective_now,
                    importer_version=nm.IMPORTER_VERSION,
                    source_file_sha256=file_sha256,
                    mapping_version=nm.MAPPING_VERSION,
                    licence_status=nm.LICENCE_STATUS,
                    attribution=nm.ATTRIBUTION,
                    transformation_notes=notes,
                )
            )
        return records
    finally:
        workbook.close()


def _parse_nutrient_value(column_letter: str, value: Any, row_index: int) -> RawNutrientValue:
    mapping = nm.map_for_column(column_letter)

    if nm.is_blank_nutrient_value(value):
        return RawNutrientValue(
            source_column=column_letter,
            source_label=mapping.source_label,
            original_value=None,
            original_unit=mapping.source_unit,
            value_state="MISSING",
        )
    if isinstance(value, bool):
        raise WorkbookStructureError(
            f"Unexpected boolean nutrient value at {column_letter}{row_index}."
        )
    if isinstance(value, (int, float)):
        return RawNutrientValue(
            source_column=column_letter,
            source_label=mapping.source_label,
            original_value=value,
            original_unit=mapping.source_unit,
            value_state="MEASURED",
        )
    raise WorkbookStructureError(
        f"Unexpected non-numeric, non-blank nutrient value at "
        f"{column_letter}{row_index}: {value!r}. This adapter was built "
        "against a workbook with numeric-or-blank nutrient cells only (no "
        "'Tr'/'<LOD' tokens)."
    )
