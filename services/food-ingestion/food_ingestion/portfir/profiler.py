"""Read-only structural profiling of the PortFIR (INSA BDCA) workbook.

Unlike :mod:`food_ingestion.portfir.parser`, this module does not fail
closed on structural drift -- it is meant to be run *against* an unknown or
possibly-drifted workbook to describe what is actually there (sheet names,
dimensions, merged ranges, formula/hidden-content counts, per-nutrient
missing-value counts, duplicate ``Cod`` detection, distinct food-group
counts) so a human can decide what to do next. It never mutates or saves the
workbook.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from . import nutrient_map as nm
from .parser import coerce_cod, sha256_file


def profile_workbook(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    file_sha256 = sha256_file(path)
    size_bytes = path.stat().st_size

    # Opened without read_only so merged_cells/hidden-state/dimensions are
    # available. Never saved -- read-only in effect.
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        sheets: dict[str, Any] = {}
        total_formulas = 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_profile = _profile_sheet(worksheet)
            total_formulas += sheet_profile["formula_count"]
            sheets[sheet_name] = sheet_profile

        defined_names = workbook.defined_names
        named_ranges = list(defined_names.keys()) if hasattr(defined_names, "keys") else list(defined_names)

        report: dict[str, Any] = {
            "path": str(path),
            "sha256": file_sha256,
            "size_bytes": size_bytes,
            "sheet_names": list(workbook.sheetnames),
            "sheets": sheets,
            "formula_count": total_formulas,
            "named_ranges": named_ranges,
        }

        if nm.DATA_SHEET_NAME in workbook.sheetnames:
            report["data_sheet"] = _profile_data_sheet(workbook[nm.DATA_SHEET_NAME])

        return report
    finally:
        workbook.close()


def _profile_sheet(worksheet: Any) -> dict[str, Any]:
    hidden_columns = [key for key, dim in worksheet.column_dimensions.items() if dim.hidden]
    hidden_rows = [key for key, dim in worksheet.row_dimensions.items() if dim.hidden]

    formula_count = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1

    return {
        "dimensions": worksheet.dimensions,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "sheet_state": worksheet.sheet_state,
        "merged_ranges": [str(rng) for rng in worksheet.merged_cells.ranges],
        "hidden_columns": hidden_columns,
        "hidden_rows": hidden_rows,
        "formula_count": formula_count,
    }


def _profile_data_sheet(worksheet: Any) -> dict[str, Any]:
    header_row_index = nm.HEADER_ROW_INDEX
    max_column = worksheet.max_column or 0
    header = {
        column_index: worksheet.cell(header_row_index, column_index).value
        for column_index in range(1, max_column + 1)
    }

    row_count = 0
    cod_seen: set[str] = set()
    duplicate_cods: dict[str, int] = {}
    missing_counts: dict[str, int] = {mapping.canonical_code: 0 for mapping in nm.NUTRIENT_MAP.values()}
    measured_counts: dict[str, int] = {mapping.canonical_code: 0 for mapping in nm.NUTRIENT_MAP.values()}
    level1_counts: dict[str, int] = {}
    level2_counts: dict[str, int] = {}
    level3_counts: dict[str, int] = {}

    max_row = worksheet.max_row or 0
    for row_index in range(nm.FIRST_DATA_ROW_INDEX, max_row + 1):
        cod_value = worksheet.cell(row_index, 1).value
        name_value = worksheet.cell(row_index, 2).value
        if cod_value is None and name_value is None:
            continue
        row_count += 1

        try:
            cod_key = coerce_cod(cod_value)
        except ValueError:
            cod_key = f"<unparseable:{cod_value!r}>"
        if cod_key in cod_seen:
            duplicate_cods[cod_key] = duplicate_cods.get(cod_key, 1) + 1
        else:
            cod_seen.add(cod_key)

        level1 = worksheet.cell(row_index, 3).value
        level2 = worksheet.cell(row_index, 4).value
        level3 = worksheet.cell(row_index, 5).value
        if level1:
            level1_counts[level1] = level1_counts.get(level1, 0) + 1
        if level2:
            level2_counts[level2] = level2_counts.get(level2, 0) + 1
        if level3:
            level3_counts[level3] = level3_counts.get(level3, 0) + 1

        for column_letter, mapping in nm.NUTRIENT_MAP.items():
            column_index = nm._column_letter_to_index(column_letter)
            value = worksheet.cell(row_index, column_index).value
            if nm.is_blank_nutrient_value(value):
                missing_counts[mapping.canonical_code] += 1
            else:
                measured_counts[mapping.canonical_code] += 1

    return {
        "header_row_index": header_row_index,
        "header": header,
        "data_row_count": row_count,
        "nutrient_column_count": len(nm.NUTRIENT_MAP),
        "distinct_cod_count": len(cod_seen),
        "duplicate_cods": duplicate_cods,
        "missing_value_counts": missing_counts,
        "measured_value_counts": measured_counts,
        "distinct_level1_count": len(level1_counts),
        "distinct_level2_count": len(level2_counts),
        "distinct_level3_count": len(level3_counts),
        "level1_counts": level1_counts,
    }
